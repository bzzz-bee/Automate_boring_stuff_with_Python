#Program to create N x N multiplication table in a spreadsheet.


import sys
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

number = int(input("Enter a number: "))
cells = number + 1

wb = openpyxl.Workbook()
sheet = wb.active

bold_labels = Font(bold=True)

while cells > 1:
    # Outer cells value given here
    sheet.cell(row=cells, column=1).value = cells - 1
    sheet.cell(row=1, column=cells).value = cells - 1
    # Code to make labels bold in spreadsheet
    sheet.cell(row=cells, column=1).font = bold_labels
    sheet.cell(row=1, column=cells).font = bold_labels

    cells -= 1

# Populate the table column with correct formula
column_length = number + 1

count = 0
while count < number:
    col_letter = get_column_letter(sheet.max_column - count)

    while column_length > 1:
        sheet[col_letter + str(column_length)] = ('=SUM(' + col_letter + '1*A'
                                               + str(column_length) + ')')

        column_length -= 1

    column_length = number + 1
    count += 1

wb.save('multiplication_table2.xlsx')