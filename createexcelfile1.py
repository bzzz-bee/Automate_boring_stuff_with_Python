
from openpyxl import Workbook
import openpyxl
wb = openpyxl.load_workbook('Exceltest1.xlsx')
sheet = wb['test1']


for row in range(2, sheet.max_row + 1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value


wb = Workbook()
wb['Sheet'].title = "test1"
sh1 = wb.active
sh1['A1'].value = "Name"
sh1['B1'].value = "Age"
sh1['C1'].value = "Category"
sh1['A2'].value = "John"
sh1['B2'].value = "15"
sh1['C2'].value = "Teen"
sh1['A3'].value = "Gina"
sh1['B3'].value = "30"
sh1['C3'].value = "Adult"
wb.save("Excel_test01.xlsx")